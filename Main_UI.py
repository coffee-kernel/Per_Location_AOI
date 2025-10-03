import tkinter as tk
import customtkinter as ctk
import time
from datetime import datetime
import os
import threading
import logging
from PIL import Image, ImageTk
from tkinter import messagebox, filedialog
import json
import pyodbc
import cv2
import numpy as np
import glob
import re
import io
import subprocess
import ftplib
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from Modules.Capture_UI import CameraApp
from Modules.Show_video import CameraDisplay

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

"""Global variables used for functions"""
MB_position = None # for inspection
img_cv = None # make the revised image in global
cap = None

"""Global function to access the TP and PART NUMBER"""
def global_function_tp_pn(main_app):
    if main_app.tp and main_app.part_number:
        print(f"Global access - TP: {main_app.tp}, P/N: {main_app.part_number}")
    else:
        print("No serial information available.")
        return

"""Class for image display from watcher"""
class ImageWatcher:
    def __init__(self, main_app, folder_path, canvas, app, update_interval=0.1):
        self.folder_path = folder_path
        self.canvas = canvas
        self.app = app
        self.update_interval = update_interval
        self.running = False
        self.latest_file = None
        self.root = app.root
        self.main_app = main_app

    def start(self):
        """Start watching the folder in a thread"""
        if not os.path.exists(self.folder_path):
            logger.error(f"Folder does not exist: {self.folder_path}")
            messagebox.showwarning("Warning", f"Image folder {self.folder_path} does not exist.")
            return
        
        self.running = True
        thread = threading.Thread(target=self._watch_folder, daemon=True)
        thread.start()

    def stop(self):
        """Stop watching the folder"""
        self.running = False

    def _watch_folder(self):
        """Continuously watch the folder for new images"""
        logger.info(f"Watching folder: {self.folder_path}")
        while self.running:
            try:
                files = [f for f in os.listdir(self.folder_path) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
                if files:
                    files = sorted(files, key=lambda f: os.path.getmtime(os.path.join(self.folder_path, f)))
                    latest_file = os.path.join(self.folder_path, files[-1])

                    if latest_file != self.latest_file:
                        self.latest_file = latest_file
                        self.root.after(0, lambda: self._update_canvas_with_image(latest_file))

                else:
                    logger.debug(f"No images found in {self.folder_path}")

                time.sleep(self.update_interval)
            except Exception as e:
                logger.error(f"Watcher error: {e}")
                time.sleep(self.update_interval)

    def _update_canvas_with_image(self, filepath):
        """Load and display the new image in the canvas with aspect ratio preservation"""
        global img_cv
        try:
            self.app.original_image = Image.open(filepath)
            logger.info(f"Loaded image {filepath} for TP: {self.main_app.tp}, P/N: {self.main_app.part_number}")

            # Get canvas dimensions
            canvas_w = self.canvas.winfo_width()
            canvas_h = self.canvas.winfo_height()

            if canvas_w <= 1 or canvas_h <= 1:
                canvas_w = self.app.root.winfo_screenwidth() - 40
                canvas_h = self.app.root.winfo_screenheight() - 140
                logger.debug(f"Using fallback canvas dimensions: {canvas_w}x{canvas_h}")

            # Get original image dimensions
            original_width, original_height = self.app.original_image.size
            canvas_ratio = canvas_w / canvas_h
            image_ratio = original_width / original_height

            # Resize image while preserving aspect ratio
            if image_ratio > canvas_ratio:
                new_width = canvas_w
                new_height = int(canvas_w / image_ratio)
            else:
                new_height = canvas_h
                new_width = int(canvas_h * image_ratio)

            img = self.app.original_image.copy()
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)

            # Saved resized image to a folder
            save_path = f"/home/nvidia/SHARPEYE_DATA/Resize_images/resized_{self.main_app.tp}_{self.main_app.part_number}.png"
            img_np = np.array(img)
            img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
            success = cv2.imwrite(save_path, img_cv)

            if success:
                logger.info(f"Resized image saved as {save_path}")
            else:
                logger.info(f"Failed to save resize image as {save_path}")

            # Store displayed dimensions and offsets
            self.app.displayed_image_width = new_width
            self.app.displayed_image_height = new_height
            self.app.image_offset_x = (canvas_w - new_width) // 2
            self.app.image_offset_y = (canvas_h - new_height) // 2

            # Clear previous content
            self.canvas.delete("all")
            self.app.drawn_rectangles.clear()

            # Display image on canvas, centered
            self.canvas.create_image(self.app.image_offset_x, self.app.image_offset_y, anchor="nw", image=tk_img)
            self.canvas.image = tk_img  # Keep reference to avoid garbage collection

            logger.info(f"Displayed image: original={original_width}x{original_height}, "
                        f"resized={new_width}x{new_height}, offset=({self.app.image_offset_x}, {self.app.image_offset_y}), "
                        f"TP: {self.main_app.tp}, P/N: {self.main_app.part_number}")
        except Exception as e:
            logger.error(f"Error displaying image: {e}")
            messagebox.showerror("Error", f"Failed to display image: {e}")

def start_drawing(event, app):
    """Record the starting point for rectangle drawing"""
    app.start_x = event.x
    app.start_y = event.y
    app.current_rectangle = app.canvas.create_rectangle(
        app.start_x, app.start_y, event.x, event.y, outline="#00FF00", width=3
    )

def update_rectangle(event, app):
    """Update the rectangle dynamically as the mouse moves"""
    if app.current_rectangle is not None:
        app.canvas.coords(
            app.current_rectangle,
            app.start_x, app.start_y,
            event.x, event.y
        )

def finish_drawing(event, app):
    """Finalize the rectangle drawing: save its coordinates to roi_list"""
    if app.current_rectangle is not None:
        coords = app.canvas.coords(app.current_rectangle)
        x1, y1, x2, y2 = coords
        x = min(x1, x2)
        y = min(y1, y2)
        width = abs(x2 - x1)
        height = abs(y2 - y1)

        # Get serial number
        serial_number = app.serial_entry.get().strip()
        if not serial_number:
            logger.error("Serial number is empty. Cannot save ROI coordinates.")
            messagebox.showwarning("Warning", "Please enter a serial number before saving ROI.")
            app.canvas.delete(app.current_rectangle)
            app.current_rectangle = None
            return

        # Adjust canvas coordinates to the displayed image's coordinate system
        displayed_width = getattr(app, "displayed_image_width", app.canvas.winfo_width())
        displayed_height = getattr(app, "displayed_image_height", app.canvas.winfo_height())

        if displayed_width <= 1 or displayed_height <= 1:
            displayed_width = app.root.winfo_screenwidth() - 40
            displayed_height = app.root.winfo_screenheight() - 140
            logger.debug(f"Using fallback dimensions: {displayed_width}x{displayed_height}")

        offset_x = getattr(app, "image_offset_x", 0)
        offset_y = getattr(app, "image_offset_y", 0)

        image_x = x - offset_x
        image_y = y - offset_y

        # Ensure coordinates are within bounds of the displayed image
        image_x = max(0, min(image_x, displayed_width - 1))
        image_y = max(0, min(image_y, displayed_height - 1))
        width = min(width, displayed_width - image_x)
        height = min(height, displayed_height - image_y)

        # Save ROI coordinates
        roi = {
            "x": x,
            "y": y,
            "width": width,
            "height": height,
            "image_x": image_x,
            "image_y": image_y,
            "serial_number": serial_number,
            "TP": app.tp if app.tp else "N/A",
            "P/N": app.part_number if app.part_number else "N/A"
        }

        app.roi_list = getattr(app, "roi_list", [])
        app.roi_list.append(roi)

        logger.info(f"Rectangle finalized: canvas coords: x={x}, y={y}, width={width}, height={height}, "
                    f"image coords: image_x={image_x}, image_y={image_y}, "
                    f"TP: {app.tp}, P/N: {app.part_number}")

        app.drawn_rectangles.append(app.current_rectangle)
        app.current_rectangle = None

def handle_clear(canvas, serial_entry, app):
    """Save each ROI as a separate image and annotation file, then clear the canvas"""
    global MB_position
    global img_cv
    
    serial_number = serial_entry.get().strip()
    if not serial_number:
        logger.error("Serial number is empty. Cannot save annotations.")
        messagebox.showwarning("Warning", "Please enter a serial number before saving annotations.")
        return

    if not hasattr(app, "roi_list") or not app.roi_list:
        logger.warning("No ROIs to save.")
        messagebox.showwarning("Warning", "No ROIs drawn to save.")
        return

    try:
        # Create save directories
        base_directory = "/home/nvidia/SHARPEYE_DATA/Profiling_data/GOOD"
        save_dir = f"{base_directory}/{app.tp}_{app.part_number}/{MB_position}" if app.tp and app.part_number else "/home/nvidia/SHARPEYE_DATA/Profiling_data/GOOD/None_None"
        os.makedirs(save_dir, exist_ok=True)

        # create directory and save resized image
        board_dict = "/home/nvidia/SHARPEYE_DATA/Profiling_data/GOOD"
        save_board_image = f"{board_dict}/{app.tp}_{app.part_number}/{MB_position}/Board_image"
        os.makedirs(save_board_image, exist_ok=True)

        # Base filename for images and JSON files
        base_filename = f"{app.tp}_{app.part_number}"

        # Get existing indices for JSON and image files to avoid conflicts
        existing_json_files = glob.glob(os.path.join(save_dir, f"{base_filename}_*.json"))
        json_indices = [
            int(re.search(r'_(\d+)\.json$', f).group(1))
            for f in existing_json_files
            if re.search(r'_(\d+)\.json$', f)
        ]
        next_json_index = max(json_indices) + 1 if json_indices else 1

        existing_image_files = glob.glob(os.path.join(save_dir, f"{base_filename}_*.png"))
        image_indices = [
            int(re.search(r'_(\d+)\.png$', f).group(1))
            for f in existing_image_files
            if re.search(r'_(\d+)\.png$', f)
        ]
        next_image_index = max(image_indices) + 1 if image_indices else 1

        existing_board_image_files = glob.glob(os.path.join(save_board_image, f"{base_filename}_*.png"))
        board_image_indices = [
            int(re.search(r'_(\d+)\.png$', f).group(1))
            for f in existing_board_image_files
            if re.search(r'_(\d+)\.png$', f)
        ]
        next_board_image_index = max(board_image_indices) + 1 if board_image_indices else 1

        # Prepare image for cropping
        if hasattr(app, "original_image") and app.original_image is not None:
            displayed_width = getattr(app, "displayed_image_width", app.canvas.winfo_width())
            displayed_height = getattr(app, "displayed_image_height", app.canvas.winfo_height())

            if displayed_width <= 1 or displayed_height <= 1:
                displayed_width = app.root.winfo_screenwidth() - 40
                displayed_height = app.root.winfo_screenheight() - 140
                logger.debug(f"Using fallback dimensions: {displayed_width}x{displayed_height}")

            # Create resized image
            original_width, original_height = app.original_image.size
            if displayed_width > 0 and displayed_height > 0:
                img = app.original_image.copy()
                img = img.resize((displayed_width, displayed_height), Image.Resampling.LANCZOS)
            else:
                logger.error("Invalid displayed dimensions. Cannot crop ROIs.")
                raise ValueError("Invalid displayed dimensions.")
        else:
            logger.error("Original image not found or not set. Cannot save ROIs.")
            raise ValueError("Original image not found or not set.")

        # Save the full board image once, independently indexed
        board_image_filename = f"{base_filename}_{next_board_image_index}.png"
        save_board_path = os.path.join(save_board_image, board_image_filename)
        img_np = np.array(img)
        img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        board_success = cv2.imwrite(save_board_path, img_cv)
        if not board_success:
            logger.error(f"Failed to save board image as {save_board_path}")
            raise ValueError(f"Failed to save board image as {save_board_path}")
        logger.info(f"Board image saved as {save_board_path}")

        # Process each ROI individually (ROI and JSON indices synced, independent of board)
        for i, roi in enumerate(app.roi_list):
            # Assign unique index for this ROI (synchronize JSON and image indices only)
            current_index = max(next_json_index + i, next_image_index + i)

            # Save ROI image
            roi_filename = f"{base_filename}_{current_index}.png"
            save_path = os.path.join(save_dir, roi_filename)

            # Crop ROI from the resized image
            image_x = roi["image_x"]
            image_y = roi["image_y"]
            width = roi["width"]
            height = roi["height"]

            roi_image = img.crop((image_x, image_y, image_x + width, image_y + height))

            # Convert Pillow image to OpenCV format for saving
            roi_image_np = np.array(roi_image)
            roi_image_cv = cv2.cvtColor(roi_image_np, cv2.COLOR_RGB2BGR)
            success = cv2.imwrite(save_path, roi_image_cv)
            if not success:
                logger.error(f"Failed to save ROI as {save_path}")
                raise ValueError(f"Failed to save ROI as {save_path}")

            logger.info(f"ROI saved as {save_path}")
            logger.debug(f"Displayed image size: {displayed_width}x{displayed_height}")
            logger.debug(f"Image coords: image_x={image_x}, image_y={image_y}, width={width}, height={height}")

            # Update ROI with filename and save path
            roi["ROI_image_index"] = roi_filename
            roi["ROI_save_path"] = save_path

            # Save ROI coordinates to a separate JSON file
            json_filename = f"{base_filename}_{current_index}.json"
            json_save_path = os.path.join(save_dir, json_filename)
            with open(json_save_path, 'w') as f:
                json.dump([roi], f, indent=4)  # Save as a single-item list
            logger.info(f"ROI coordinates saved to {json_save_path}")

        # Update indices for next save operation
        next_json_index = current_index + 1
        next_image_index = current_index + 1
        next_board_image_index += 1  # Increment once since board is saved once

        # Clear the canvas and reset lists
        canvas.delete("all")
        app.drawn_rectangles.clear()
        app.roi_list = []
        messagebox.showinfo("Info", f"Saved profiling successfully!")
        time.sleep(0.5)
    except Exception as e:
        logger.error(f"Error saving ROI coordinates or images: {e}")
        messagebox.showerror("Error", f"Failed to save ROI coordinates or images: {e}")
        return


# ***************************************************** FOR NG PROFILING ************************************************************* #

def start_drawing_NG(event, app):
    """Record the starting point for rectangle drawing"""
    app.start_x = event.x
    app.start_y = event.y
    app.current_rectangle = app.canvas.create_rectangle(
        app.start_x, app.start_y, event.x, event.y, outline="red", width=3
    )

def update_rectangle_NG(event, app):
    """Update the rectangle dynamically as the mouse moves"""
    if app.current_rectangle is not None:
        app.canvas.coords(
            app.current_rectangle,
            app.start_x, app.start_y,
            event.x, event.y
        )

def finish_drawing_NG(event, app):
    """Finalize the rectangle drawing: save its coordinates to roi_list"""
    if app.current_rectangle is not None:
        coords = app.canvas.coords(app.current_rectangle)
        x1, y1, x2, y2 = coords
        x = min(x1, x2)
        y = min(y1, y2)
        width = abs(x2 - x1)
        height = abs(y2 - y1)

        # Get serial number
        serial_number = app.serial_entry.get().strip()
        if not serial_number:
            logger.error("Serial number is empty. Cannot save ROI coordinates.")
            messagebox.showwarning("Warning", "Please enter a serial number before saving ROI.")
            app.canvas.delete(app.current_rectangle)
            app.current_rectangle = None
            return

        # Adjust canvas coordinates to the displayed image's coordinate system
        displayed_width = getattr(app, "displayed_image_width", app.canvas.winfo_width())
        displayed_height = getattr(app, "displayed_image_height", app.canvas.winfo_height())

        if displayed_width <= 1 or displayed_height <= 1:
            displayed_width = app.root.winfo_screenwidth() - 40
            displayed_height = app.root.winfo_screenheight() - 140
            logger.debug(f"Using fallback dimensions: {displayed_width}x{displayed_height}")

        offset_x = getattr(app, "image_offset_x", 0)
        offset_y = getattr(app, "image_offset_y", 0)

        image_x = x - offset_x
        image_y = y - offset_y

        # Ensure coordinates are within bounds of the displayed image
        image_x = max(0, min(image_x, displayed_width - 1))
        image_y = max(0, min(image_y, displayed_height - 1))
        width = min(width, displayed_width - image_x)
        height = min(height, displayed_height - image_y)

        # Save ROI coordinates
        roi = {
            "x": x,
            "y": y,
            "width": width,
            "height": height,
            "image_x": image_x,
            "image_y": image_y,
            "serial_number": serial_number,
            "TP": app.tp if app.tp else "N/A",
            "P/N": app.part_number if app.part_number else "N/A"
        }

        app.roi_list = getattr(app, "roi_list", [])
        app.roi_list.append(roi)

        logger.info(f"Rectangle finalized: canvas coords: x={x}, y={y}, width={width}, height={height}, "
                    f"image coords: image_x={image_x}, image_y={image_y}, "
                    f"TP: {app.tp}, P/N: {app.part_number}")

        app.drawn_rectangles.append(app.current_rectangle)
        app.current_rectangle = None

def handle_clear_NG(canvas, serial_entry, app):
    """Save each ROI as a separate image and annotation file, then clear the canvas"""
    global MB_position

    serial_number = serial_entry.get().strip()
    if not serial_number:
        logger.error("Serial number is empty. Cannot save annotations.")
        messagebox.showwarning("Warning", "Please enter a serial number before saving annotations.")
        return

    if not hasattr(app, "roi_list") or not app.roi_list:
        logger.warning("No ROIs to save.")
        messagebox.showwarning("Warning", "No ROIs drawn to save.")
        return

    try:
        # Create save directories
        base_directory = "/home/nvidia/SHARPEYE_DATA/Profiling_data_NG/NG"
        save_dir = f"{base_directory}/{app.tp}_{app.part_number}/{MB_position}" if app.tp and app.part_number else "/home/nvidia/SHARPEYE_DATA/Profiling_data_NG/NG/None_None"
        os.makedirs(save_dir, exist_ok=True)

        # create directory and save resized image
        board_dict = "/home/nvidia/SHARPEYE_DATA/Profiling_data_NG/NG"
        save_board_image = f"{board_dict}/{app.tp}_{app.part_number}/{MB_position}/Board_image"
        os.makedirs(save_board_image, exist_ok=True)

        # Base filename for images and JSON files
        base_filename = f"{app.tp}_{app.part_number}"

        # Get existing indices for JSON and image files to avoid conflicts
        existing_json_files = glob.glob(os.path.join(save_dir, f"{base_filename}_*.json"))
        json_indices = [
            int(re.search(r'_(\d+)\.json$', f).group(1))
            for f in existing_json_files
            if re.search(r'_(\d+)\.json$', f)
        ]
        next_json_index = max(json_indices) + 1 if json_indices else 1

        existing_image_files = glob.glob(os.path.join(save_dir, f"{base_filename}_*.png"))
        image_indices = [
            int(re.search(r'_(\d+)\.png$', f).group(1))
            for f in existing_image_files
            if re.search(r'_(\d+)\.png$', f)
        ]
        next_image_index = max(image_indices) + 1 if image_indices else 1

        existing_board_image_files = glob.glob(os.path.join(save_board_image, f"{base_filename}_*.png"))
        board_image_indices = [
            int(re.search(r'_(\d+)\.png$', f).group(1))
            for f in existing_board_image_files
            if re.search(r'_(\d+)\.png$', f)
        ]
        next_board_image_index = max(board_image_indices) + 1 if board_image_indices else 1

        # Prepare image for cropping
        if hasattr(app, "original_image") and app.original_image is not None:
            displayed_width = getattr(app, "displayed_image_width", app.canvas.winfo_width())
            displayed_height = getattr(app, "displayed_image_height", app.canvas.winfo_height())

            if displayed_width <= 1 or displayed_height <= 1:
                displayed_width = app.root.winfo_screenwidth() - 40
                displayed_height = app.root.winfo_screenheight() - 140
                logger.debug(f"Using fallback dimensions: {displayed_width}x{displayed_height}")

            # Create resized image
            original_width, original_height = app.original_image.size
            if displayed_width > 0 and displayed_height > 0:
                img = app.original_image.copy()
                img = img.resize((displayed_width, displayed_height), Image.Resampling.LANCZOS)
            else:
                logger.error("Invalid displayed dimensions. Cannot crop ROIs.")
                raise ValueError("Invalid displayed dimensions.")
        else:
            logger.error("Original image not found or not set. Cannot save ROIs.")
            raise ValueError("Original image not found or not set.")

        # Save the full board image once, independently indexed
        board_image_filename = f"{base_filename}_{next_board_image_index}.png"
        save_board_path = os.path.join(save_board_image, board_image_filename)
        img_np = np.array(img)
        img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        board_success = cv2.imwrite(save_board_path, img_cv)
        if not board_success:
            logger.error(f"Failed to save board image as {save_board_path}")
            raise ValueError(f"Failed to save board image as {save_board_path}")
        logger.info(f"Board image saved as {save_board_path}")

        # Process each ROI individually (ROI and JSON indices synced, independent of board)
        for i, roi in enumerate(app.roi_list):
            # Assign unique index for this ROI (synchronize JSON and image indices only)
            current_index = max(next_json_index + i, next_image_index + i)

            # Save ROI image
            roi_filename = f"{base_filename}_{current_index}.png"
            save_path = os.path.join(save_dir, roi_filename)

            # Crop ROI from the resized image
            image_x = roi["image_x"]
            image_y = roi["image_y"]
            width = roi["width"]
            height = roi["height"]

            roi_image = img.crop((image_x, image_y, image_x + width, image_y + height))

            # Convert Pillow image to OpenCV format for saving
            roi_image_np = np.array(roi_image)
            roi_image_cv = cv2.cvtColor(roi_image_np, cv2.COLOR_RGB2BGR)
            success = cv2.imwrite(save_path, roi_image_cv)
            if not success:
                logger.error(f"Failed to save ROI as {save_path}")
                raise ValueError(f"Failed to save ROI as {save_path}")

            logger.info(f"ROI saved as {save_path}")
            logger.debug(f"Displayed image size: {displayed_width}x{displayed_height}")
            logger.debug(f"Image coords: image_x={image_x}, image_y={image_y}, width={width}, height={height}")

            # Update ROI with filename and save path
            roi["ROI_image_index"] = roi_filename
            roi["ROI_save_path"] = save_path

            # Save ROI coordinates to a separate JSON file
            json_filename = f"{base_filename}_{current_index}.json"
            json_save_path = os.path.join(save_dir, json_filename)
            with open(json_save_path, 'w') as f:
                json.dump([roi], f, indent=4)  # Save as a single-item list
            logger.info(f"ROI coordinates saved to {json_save_path}")

        # Update indices for next save operation
        next_json_index = current_index + 1
        next_image_index = current_index + 1
        next_board_image_index += 1  # Increment once since board is saved once

        # Clear the canvas and reset lists
        canvas.delete("all")
        app.drawn_rectangles.clear()
        app.roi_list = []
        messagebox.showinfo("Info", f"Saved NG profiling successfully!")
        time.sleep(0.5)
    except Exception as e:
        logger.error(f"Error saving ROI coordinates or images: {e}")
        messagebox.showerror("Error", f"Failed to save ROI coordinates or images: {e}")
        return
    
    
# ***************************************************** MAIN APPLICATION ************************************************************* #

class MainApplication:
    def __init__(self, root):
        # CustomTkinter settings
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")

        # Main window setup
        self.root = root
        self.root.title("SharpEye AOI NB")
        self.root.configure(bg="black")

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")

        # Top Frame
        self.top_frame = ctk.CTkFrame(self.root, corner_radius=0, fg_color="#1e1e1e")
        self.top_frame.pack(side="top", fill="x")

        # Serial Entry
        self.label = ctk.CTkLabel(self.top_frame, text="SERIAL NUMBER:", font=("Arial", 16, "bold"))
        self.label.pack(side="left", padx=10, pady=15)
        self.serial_entry = ctk.CTkEntry(self.top_frame, width=350, font=("Arial", 16))
        self.serial_entry.pack(side="left", padx=5, pady=15)

        # Canvas
        self.canvas_frame = ctk.CTkFrame(self.root, corner_radius=0, fg_color="#222")
        self.canvas_frame.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(self.canvas_frame, bg="black", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()

        # Variables for drawing and image display
        self.start_x, self.start_y = None, None
        self.current_rectangle = None
        self.drawn_rectangles = []
        self.roi_list = []
        self.original_image = None
        self.displayed_image_width = 0
        self.displayed_image_height = 0
        self.image_offset_x = 0
        self.image_offset_y = 0
        self.serial_entry.focus_set()
        
        # For model and partnumber variables
        self.tp = None
        self.part_number = None

        # Initialize the camera app from module
        self.capture = CameraApp()

        # Initialize the camera display from module
        self.camera_display = CameraDisplay(self.root, self.canvas, self.screen_height, self.screen_width)

        # Initialize ImageWatcher with correct folder path
        folder_path = "/home/nvidia/SHARPEYE_DATA/Captured_images"
        self.watcher = ImageWatcher(self, folder_path, self.canvas, self, update_interval=0.1)

        # Initialize camera
        self.capture.initialize_camera()

        # Bind uppercase key release
        self.serial_entry.bind("<KeyRelease>", lambda event: self.to_uppercase(event))
        self.serial_entry.bind("<KeyRelease-Return>", lambda event: self.handle_enter_key(event))

        # Add closing protocol
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Dropdown selection
        self.dropdown = ctk.CTkComboBox(
            self.top_frame, values=["Inspection Process", "Profiling Process", "Profiling Process NG"], width=174, dropdown_fg_color="#333", command=lambda choice: self.dropdown_choice(choice), state="readonly"
        )
        self.dropdown.pack(side="left", padx=5, pady=15)

        # default value
        self.dropdown.set("---Select Process---")

        # selection of TOP VIEW and BOTTOM VIEW 
        """FOR INSPECTION"""
        self.dropdown_mb_position_inspect = ctk.CTkComboBox(
            self.top_frame, values=["TOP VIEW", "BOTTOM VIEW"], width=186, dropdown_fg_color="#333", command=self.on_mb_position_inspect, state="readonly"
        )
        self.dropdown_mb_position_inspect.pack_forget()

        # default value
        self.dropdown_mb_position_inspect.set("---Select MB Position---")

        # selection of TOP VIEW and BOTTOM VIEW 
        """FOR PROFILING"""
        self.dropdown_mb_position_profile = ctk.CTkComboBox(
            self.top_frame, values=["TOP VIEW", "BOTTOM VIEW"], width=186, dropdown_fg_color="#333", command=self.on_mb_position_profile, state="readonly"
        )
        self.dropdown_mb_position_profile.pack_forget()

        # default value
        self.dropdown_mb_position_profile.set("---Select MB Position---")

        """FOR NG PROFILING"""
        self.dropdown_mb_position_profile_NG = ctk.CTkComboBox(
            self.top_frame, values=["TOP VIEW", "BOTTOM VIEW"], width=186, dropdown_fg_color="#333", command=self.on_mb_position_profile_NG, state="readonly"
        )
        self.dropdown_mb_position_profile_NG.pack_forget()

        # default value
        self.dropdown_mb_position_profile_NG.set("---Select MB Position---")

        # Buttons for inspection and profiling
        icon_path_capture_profile = "/home/nvidia/SHARPEYE_DATA/Assets/capture.png"
        icon_image_capture_profile = ctk.CTkImage(Image.open(icon_path_capture_profile), size=(20, 20))
        self.capture_btn_profile = ctk.CTkButton(self.top_frame, text="PROFILE REF", font=("Arial", 12, "bold"), command=self.capture_function, image=icon_image_capture_profile, compound="left")
        self.capture_btn_profile.pack_forget

        icon_path_capture_inspect = "/home/nvidia/SHARPEYE_DATA/Assets/capture.png"
        icon_image_capture_inspect = ctk.CTkImage(Image.open(icon_path_capture_inspect), size=(20, 20))
        self.capture_btn_inspect = ctk.CTkButton(self.top_frame, text="CAPTURE", font=("Arial", 12, "bold"), command=self.capture_function_inspection, image=icon_image_capture_inspect, compound="left") # For capture inspection
        self.capture_btn_inspect.pack_forget()

        icon_path_compare = "/home/nvidia/SHARPEYE_DATA/Assets/inspect.png"
        icon_image_compare = ctk.CTkImage(Image.open(icon_path_compare), size=(20, 20))
        self.compare_btn_inspect = ctk.CTkButton(self.top_frame, text="INSPECT", font=("Arial", 12, "bold"), command=self.match_and_annotate, image=icon_image_compare, compound="left")
        self.compare_btn_inspect.pack_forget()

        icon_path_save = "/home/nvidia/SHARPEYE_DATA/Assets/save.png"
        icon_image_save = ctk.CTkImage(Image.open(icon_path_save), size=(20, 20))
        self.clear_btn = ctk.CTkButton(self.top_frame, text="SAVE", width=120, font=("Arial", 12, "bold"), command=lambda: handle_clear(self.canvas, self.serial_entry, self), image=icon_image_save, compound="left")
        self.clear_btn.pack_forget()

        icon_path_reinspect = "/home/nvidia/SHARPEYE_DATA/Assets/paper.png"
        icon_image_reinspect = ctk.CTkImage(Image.open(icon_path_reinspect), size=(20, 20))
        self.re_inspect_btn = ctk.CTkButton(self.top_frame, text="RE-INSPECT", font=("Arial", 12, "bold"), command=self.show_video, image=icon_image_reinspect, compound="left")
        self.re_inspect_btn.pack_forget()

        icon_path_refresh = "/home/nvidia/SHARPEYE_DATA/Assets/clean.png"
        icon_image_refresh = ctk.CTkImage(Image.open(icon_path_refresh), size=(20, 20))
        self.clear_rect_button = ctk.CTkButton(self.top_frame, text="REFRESH", font=("Arial", 12, "bold"), fg_color="orange", hover_color="darkorange", command=self.clear_rectangles, image=icon_image_refresh, compound="left")
        self.clear_rect_button.pack_forget()

        icon_path_delete_inspect = "/home/nvidia/SHARPEYE_DATA/Assets/trash-bin.png"
        icon_image_delete_inspect = ctk.CTkImage(Image.open(icon_path_delete_inspect), size=(20, 20))
        self.clear_serial = ctk.CTkButton(self.top_frame, text="DELETE", fg_color="red", hover_color="darkred", width=120, font=("Arial", 12, "bold"), command=self.clear_serial_number, image=icon_image_delete_inspect, compound="left")
        self.clear_serial.pack_forget()

        # For profile capture toggle
        icon_path_delete_profile = "/home/nvidia/SHARPEYE_DATA/Assets/trash-bin.png"
        icon_image_delete_profile = ctk.CTkImage(Image.open(icon_path_delete_profile), size=(20, 20))
        self.clear_serial_profile = ctk.CTkButton(self.top_frame, text="DELETE", fg_color="red", hover_color="darkred", width=120, font=("Arial", 12, "bold"), command=self.clear_serial_number, image=icon_image_delete_profile, compound="left")
        self.clear_serial_profile.pack_forget()

        # For NG profiling
        icon_path_capture_profile_NG = "/home/nvidia/SHARPEYE_DATA/Assets/capture.png"
        icon_image_capture_profile_NG = ctk.CTkImage(Image.open(icon_path_capture_profile_NG), size=(20, 20))
        self.capture_btn_profile_NG = ctk.CTkButton(self.top_frame, text="PROFILE REF NG", font=("Arial", 12, "bold"), command=self.capture_function_NG, image=icon_image_capture_profile_NG, compound="left")
        self.capture_btn_profile_NG.pack_forget

        icon_path_save_NG = "/home/nvidia/SHARPEYE_DATA/Assets/save.png"
        icon_image_save_NG = ctk.CTkImage(Image.open(icon_path_save_NG), size=(20, 20))
        self.clear_btn_NG = ctk.CTkButton(self.top_frame, text="SAVE NG", width=120, font=("Arial", 12, "bold"), command=lambda: handle_clear_NG(self.canvas, self.serial_entry, self), image=icon_image_save_NG, compound="left")
        self.clear_btn_NG.pack_forget()


    # get the USB camera device
    def get_camera_usb_device(self):
        try:
            result = subprocess.run(["lsusb"], capture_output=True, text=True)
            for line in result.stdout.splitlines():
                if "Canon" in line:
                    match = re.search(r"Bus (\d+) Device (\d+)", line)
                    if match:
                        bus, dev = match.groups()
                        return f"/dev/bus/usb/{bus}/{dev}"
            return None
        except Exception as e:
            print(f"Error finding USB device: {str(e)}")
            return None
    
    # unmount usb storage
    def unmount_usb_storage(self):
        usb_device = self.get_camera_usb_device()
        if usb_device:
            subprocess.run(["udisksctl", "unmount", "-b", usb_device], capture_output=True, text=True)
        else:
            print("No USB device detected!")

    # mount the usb storage
    def mount_usb_storage(self):
        usb_device = self.get_camera_usb_device()
        if usb_device:
            subprocess.run(["udisksctl", "mount", "-b", usb_device], capture_output=True, text=True)
        else:
            print("No USB device detected!")

    def show_video(self):
        """Show the video feed."""
        self.camera_display.show_video()
        self.mount_usb_storage()

    def hide_video(self):
        """Hide the video feed."""
        self.camera_display.hide_video()

    def to_uppercase(self, event):
        """Convert serial entry to uppercase"""
        current_text = self.serial_entry.get()
        self.serial_entry.delete(0, "end")
        self.serial_entry.insert(0, current_text.upper())

    def handle_enter_key(self, event):
        """Handle Enter key press for serial number validation"""
        if event.keysym == "Return":
            serial_number = self.serial_entry.get()
            self.check_serial_exists(serial_number)
            print(f"Enter key pressed. Serial Number: {serial_number}")

    def clear_serial_number(self):
        """handle to clear serial number only"""
        self.serial_entry.delete(0, "end")
        self.dropdown_mb_position_inspect.set("---Select MB Position---")
        self.dropdown_mb_position_profile.set("---Select MB Position---")
        self.dropdown_mb_position_profile_NG.set("---Select MB Position---")

    def re_position_widgets(self):
        self.dropdown_mb_position_profile.pack(side="left", padx=5, pady=15)
        self.capture_btn_profile.pack(side="left", padx=5, pady=15)
        self.clear_btn.pack(side="left", padx=5, pady=15)
        self.clear_rect_button.pack(side="left", padx=5, pady=15)
        self.clear_serial_profile.pack(side="left", padx=5, pady=15)
        self.serial_entry.focus_set()

    def re_position_widgets_NG(self):
        self.dropdown_mb_position_profile_NG.pack(side="left", padx=5, pady=15)
        self.capture_btn_profile_NG.pack(side="left", padx=5, pady=15)
        self.clear_btn_NG.pack(side="left", padx=5, pady=15)
        self.clear_rect_button.pack(side="left", padx=5, pady=15)
        self.clear_serial_profile.pack(side="left", padx=5, pady=15)
        self.serial_entry.focus_set()

    # for dropdown function # Ongoing
    def dropdown_choice(self, choice):
        if choice == "Inspection Process":
            """Toggle to see inspection process buttons"""
            self.dropdown_mb_position_inspect.pack(side="left", padx=5, pady=15)
            self.capture_btn_inspect.pack(side="left", padx=5, pady=15)
            self.compare_btn_inspect.pack(side="left", padx=5, pady=15)
            self.re_inspect_btn.pack(side="left", padx=5, pady=15)
            self.clear_serial.pack(side="left", padx=5, pady=15)
            self.serial_entry.focus_set()

            """pack forget for profile capture"""
            self.capture_btn_profile.pack_forget()
            self.clear_btn.pack_forget()
            self.clear_rect_button.pack_forget()
            self.clear_btn_NG.pack_forget()
            self.capture_btn_profile_NG.pack_forget()
            self.clear_serial_profile.pack_forget()
            self.dropdown_mb_position_profile.pack_forget()
            self.dropdown_mb_position_profile_NG.pack_forget()
            self.capture_btn_profile_NG.pack_forget()
            self.dropdown_mb_position_profile.set("---Select MB Position---")
            self.dropdown_mb_position_profile_NG.set("---Select MB Position---")

        elif choice == "Profiling Process":
            """Toggle to see profile process buttons"""
            self.clear_rect_button.pack_forget()
            self.clear_serial_profile.pack_forget()
            self.root.after(200, lambda: self.re_position_widgets())

            """pack forget for inspection"""
            self.capture_btn_inspect.pack_forget()
            self.compare_btn_inspect.pack_forget()
            self.re_inspect_btn.pack_forget()
            self.clear_serial.pack_forget()
            self.dropdown_mb_position_inspect.pack_forget()
            self.dropdown_mb_position_profile_NG.pack_forget()
            self.clear_btn_NG.pack_forget()
            self.capture_btn_profile_NG.pack_forget()
            self.dropdown_mb_position_inspect.set("---Select MB Position---")
            self.dropdown_mb_position_profile_NG.set("---Select MB Position---")

        elif choice == "Profiling Process NG":
            """Toggle to see profile process buttons"""
            self.clear_rect_button.pack_forget()
            self.clear_serial_profile.pack_forget()
            self.root.after(200, lambda: self.re_position_widgets_NG())

            """pack forget for inspection"""
            self.capture_btn_profile.pack_forget()
            self.clear_btn.pack_forget()
            self.dropdown_mb_position_profile.pack_forget()
            self.dropdown_mb_position_profile.set("---Select MB Position---")
            self.capture_btn_inspect.pack_forget()
            self.compare_btn_inspect.pack_forget()
            self.re_inspect_btn.pack_forget()
            self.clear_serial.pack_forget()
            self.dropdown_mb_position_inspect.pack_forget()
            self.dropdown_mb_position_inspect.set("---Select MB Position---")

            """Default"""
        else: 
            self.capture_btn_profile.pack_forget()
            self.clear_btn.pack_forget()
            self.clear_rect_button.pack_forget()
            self.capture_btn_inspect.pack_forget()
            self.compare_btn_inspect.pack_forget()
            self.re_inspect_btn.pack_forget()
            self.clear_serial.pack_forget()
            self.clear_serial_profile.pack_forget()
            self.dropdown_mb_position_inspect.pack_forget()
            self.dropdown_mb_position_profile.pack_forget()
            self.dropdown_mb_position_profile_NG.pack_forget()
            self.capture_btn_profile_NG.pack_forget()
            self.clear_btn_NG.pack_forget()
            self.dropdown_mb_position_inspect.set("---Select MB Position---")
            self.dropdown_mb_position_profile.set("---Select MB Position---")
            self.dropdown_mb_position_profile_NG.set("---Select MB Position---")

    # function for selecting value on inspection and profiling process
    def on_mb_position_inspect(self, selected_value):
        global MB_position
        MB_position = selected_value
        logger.info(f"Selected MB Position: {MB_position}")

    def on_mb_position_profile(self, selected_value):
        global MB_position
        MB_position = selected_value
        logger.info(f"Selected MB Position for profiling {MB_position}")

    def on_mb_position_profile_NG(self, selected_value):
        global MB_position
        MB_position = selected_value
        logger.info(f"Selected MB Position for profiling {MB_position}")

    def check_serial_exists(self, serial_number):
        # Configure SQL Server connection details
        server = '192.168.1.152,1433\\SQLEXPRESS'
        database = 'SharpeyeWeb'
        username = "CBE_QAS"
        password = "cre100"
        trust_certificate = 'Yes'
        connection_timeout = '30'
        table_name = 'Upload_serial_number'
        column_name = 'SN'

        conn_str = (
            f"DRIVER={{ODBC Driver 18 for SQL Server}};"
            f"SERVER={server};"
            f"DATABASE={database};"
            f"UID={username};"
            f"PWD={password};"
            f"TrustServerCertificate={trust_certificate};"
            f"Connection Timeout={connection_timeout};"
        )

        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            query = f"SELECT COUNT(*), TP, [P/N] FROM [{table_name}] WHERE [{column_name}] = ? GROUP BY TP, [P/N]"
            cursor.execute(query, (serial_number,))
            result = cursor.fetchone()
            conn.close()

            if result and result[0] > 0:
                # Serial number exists, store and print TP and P/N
                self.tp = result[1] if result[1] is not None else "N/A"
                self.part_number = result[2] if result[2] is not None else "N/A"
                logger.info(f"Serial number '{serial_number}' exists in database. TP: {self.tp}, P/N: {self.part_number}")
                print(f"Serial number: {serial_number}, TP: {self.tp}, P/N: {self.part_number}")
                return False
            else:
                # Serial number does not exist
                self.tp = None
                self.part_number = None
                logger.info(f"Serial number '{serial_number}' does not exist in database. Proceeding.")
                messagebox.showwarning("Info", f"Serial number {serial_number} does not exist!")
                self.serial_entry.delete(0, 'end')
                return True
        except pyodbc.Error as e:
            logger.error(f"Database connection/check failed: {e}. Skipping serial check.")
            messagebox.showerror("Database Error", f"Failed to check serial number in database: {e}\nProceeding without check.")
            self.tp = None
            self.part_number = None
            return True

    # ******************* Display results here ********************* #

    def display_annotated_image(self, image_path):
        """Display the annotated image on the canvas with aspect ratio preservation"""
        try:
            logger.info(f"Attempting to display annotated image from {image_path}")
            
            # Verify canvas and root existence
            canvas = getattr(self, 'canvas', None)
            if canvas is None and hasattr(self, 'app'):
                canvas = getattr(self.app, 'canvas', None)
            if canvas is None:
                logger.error("Canvas not found in self or self.app")
                raise AttributeError("Canvas not found")
            
            root = getattr(self, 'root', None)
            if root is None and hasattr(self, 'app'):
                root = getattr(self.app, 'root', None)
            if root is None:
                logger.error("Root window not found in self or self.app")
                raise AttributeError("Root window not found")
            
            # Load the image with OpenCV
            annotated_img = cv2.imread(image_path)
            if annotated_img is None:
                logger.error(f"Could not load annotated image from {image_path}")
                raise ValueError(f"Could not load annotated image from {image_path}")
            
            # Convert BGR to RGB for PIL
            annotated_img_rgb = cv2.cvtColor(annotated_img, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(annotated_img_rgb)
            logger.info(f"Loaded annotated image from {image_path}, size={pil_img.size}")

            # Get canvas dimensions
            canvas_w = canvas.winfo_width()
            canvas_h = canvas.winfo_height()

            if canvas_w <= 1 or canvas_h <= 1:
                canvas_w = root.winfo_screenwidth() - 40
                canvas_h = root.winfo_screenheight() - 140
                logger.debug(f"Using fallback canvas dimensions: {canvas_w}x{canvas_h}")

            # Get original image dimensions
            original_width, original_height = pil_img.size
            canvas_ratio = canvas_w / canvas_h
            image_ratio = original_width / original_height

            # Resize image while preserving aspect ratio
            if image_ratio > canvas_ratio:
                new_width = canvas_w
                new_height = int(canvas_w / image_ratio)
            else:
                new_height = canvas_h
                new_width = int(canvas_h * image_ratio)

            # Resize the image
            pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            tk_img = ImageTk.PhotoImage(pil_img)
            logger.info(f"Resized image to {new_width}x{new_height}")

            # Store displayed dimensions and offsets
            self.displayed_image_width = new_width
            self.displayed_image_height = new_height
            self.image_offset_x = (canvas_w - new_width) // 2
            self.image_offset_y = (canvas_h - new_height) // 2

            # Clear previous content
            canvas.delete("all")
            self.drawn_rectangles = getattr(self, "drawn_rectangles", [])
            self.drawn_rectangles.clear()
            logger.info("Cleared canvas content")

            # Display image on canvas, centered
            canvas.create_image(self.image_offset_x, self.image_offset_y, anchor="nw", image=tk_img)
            self.canvas_image = tk_img  # Store reference to avoid garbage collection
            logger.info(f"Created canvas image at offset ({self.image_offset_x}, {self.image_offset_y})")

            # Update canvas to ensure rendering
            canvas.update()
            logger.info(f"Updated canvas for display, TP: {self.tp}, P/N: {self.part_number}")
        except Exception as e:
            logger.error(f"Error displaying annotated image: {e}")
            raise


    # ******************* Matching template ************************ #

    def match_and_annotate(self):
        global MB_position

        """Perform template matching for all ROI images and annotate live video feed with rectangles using template match location and JSON width/height (CUDA-accelerated)"""
        base_roi_folder = "/home/nvidia/SHARPEYE_DATA/Profiling_data"
        base_roi_ng_folder = "/home/nvidia/SHARPEYE_DATA/Profiling_data_NG"
        base_json_folder = "/home/nvidia/SHARPEYE_DATA/Profiling_data"
        base_json_ng_folder = "/home/nvidia/SHARPEYE_DATA/Profiling_data_NG"

        # confidence threshold for trial
        confidence_threshold = 0.75  # for trial

        # Database config
        driver = '{ODBC Driver 18 for SQL Server}'
        server = '192.168.1.152,1433\\SQLEXPRESS'
        database = 'SharpeyeWeb'
        username = "CBE_QAS"
        password = "cre100"
        trust_certificate = 'Yes'
        connection_timeout = '30'

        # Validate inputs
        for var_name, var_value in [("self.tp", self.tp), ("self.part_number", self.part_number), ("MB_position", MB_position)]:
            if not isinstance(var_value, str):
                logger.error(f"{var_name} is not a string: {var_value} (type: {type(var_value)})")
                raise ValueError(f"{var_name} must be a string, got {type(var_value)}: {var_value}")

        # Construct specific folder paths based on TP and P/N for GOOD and NG folders
        roi_folders = [
            (os.path.join(base_roi_folder, "GOOD", f"{self.part_number}_{self.tp}", MB_position),), 
            (os.path.join(base_roi_ng_folder, "NG", f"{self.part_number}_{self.tp}", MB_position),)
        ]
        json_folders = [
            (os.path.join(base_json_folder, "GOOD", f"{self.part_number}_{self.tp}", MB_position),),
            (os.path.join(base_json_ng_folder, "NG", f"{self.part_number}_{self.tp}", MB_position),)
        ]
        folder_types = ["GOOD", "NG"]  # Corresponding folder types

        # Debug: Log folder lists and MB_position
        logger.info(f"MB_position: {MB_position}")
        logger.info(f"roi_folders: {roi_folders}")
        logger.info(f"json_folders: {json_folders}")
        for i, (roi_folder_tuple, json_folder_tuple) in enumerate(zip(roi_folders, json_folders)):
            logger.info(f"roi_folders[{i}]: {roi_folder_tuple}, type: {type(roi_folder_tuple)}, len: {len(roi_folder_tuple)}")
            logger.info(f"json_folders[{i}]: {json_folder_tuple}, type: {type(json_folder_tuple)}, len: {len(json_folder_tuple)}")

        # Initialize VideoCapture using GStreamer pipeline if not already done
        if not hasattr(self, 'cap') or self.cap is None:
            if not self.capture.initialize_camera():
                logger.error("Failed to initialize capture with GStreamer pipeline")
                raise ValueError("Failed to initialize capture with GStreamer pipeline")

        # Get initial frame for dimensions
        ret, frame = self.cap.read()
        if not ret:
            logger.error("Could not read initial frame")
            raise ValueError("Could not read initial frame")
        img_height, img_width = frame.shape[:2]
        logger.info(f"Live frame dimensions: {img_width}x{img_height}")

        # Get the GOOD ROI folder for reference images
        good_roi_folder = f"/home/nvidia/SHARPEYE_DATA/Profiling_data/GOOD/{self.tp}_{self.part_number}/{MB_position}/Board_image"

        # One-time auto-alignment setup (compute reference and homography once using initial frame)
        reference_img = None
        homography_matrix = None
        initial_frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gpu_initial = cv2.cuda_GpuMat()
        gpu_initial.upload(frame)
        gpu_initial_gray = cv2.cuda.cvtColor(gpu_initial, cv2.COLOR_BGR2GRAY)

        if os.path.exists(good_roi_folder):
            png_files = [f for f in os.listdir(good_roi_folder) if f.lower().endswith('.png')]
            logger.info(f"Found {len(png_files)} potential reference images in {good_roi_folder}")

            orb_gpu = cv2.cuda.ORB_create(maxFeatures=5000)
            kp2_gpu, des2_gpu = orb_gpu.detectAndCompute(gpu_initial_gray, None)

            if des2_gpu is not None:
                des2 = des2_gpu.download()
                best_reference_path = None
                best_score = -1
                for png_file in png_files:
                    candidate_path = os.path.join(good_roi_folder, png_file)
                    candidate_img = cv2.imread(candidate_path)
                    if candidate_img is None:
                        continue

                    candidate_gray = cv2.cvtColor(candidate_img, cv2.COLOR_BGR2GRAY)
                    cand_height, cand_width = candidate_gray.shape
                    if cand_width > img_width or cand_height > img_height:
                        logger.info(f"Skipping candidate {png_file} as it is larger than frame")
                        continue

                    gpu_candidate = cv2.cuda_GpuMat()
                    gpu_candidate.upload(candidate_img)
                    gpu_candidate_gray = cv2.cuda.cvtColor(gpu_candidate, cv2.COLOR_BGR2GRAY)

                    kp1_gpu, des1_gpu = orb_gpu.detectAndCompute(gpu_candidate_gray, None)
                    if des1_gpu is None:
                        gpu_candidate.release()
                        continue
                    des1 = des1_gpu.download()

                    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
                    matches = bf.match(des1, des2)
                    matches = sorted(matches, key=lambda x: x.distance)
                    good_matches = [m for m in matches[:50] if m.distance < 50]
                    score = len(good_matches)

                    logger.info(f"Candidate {png_file}: {score} good matches")

                    if score > best_score:
                        best_score = score
                        best_reference_path = candidate_path

                    gpu_candidate.release()

                if best_reference_path is not None:
                    logger.info(f"Selected best reference image: {os.path.basename(best_reference_path)} with {best_score} good matches")
                    reference_img = cv2.imread(best_reference_path)

        # Compute homography once if reference found
        if reference_img is not None:
            try:
                gpu_reference = cv2.cuda_GpuMat()
                gpu_reference.upload(reference_img)
                gpu_reference_gray = cv2.cuda.cvtColor(gpu_reference, cv2.COLOR_BGR2GRAY)

                orb_gpu = cv2.cuda.ORB_create(maxFeatures=5000)
                kp1_gpu, des1_gpu = orb_gpu.detectAndCompute(gpu_reference_gray, None)
                kp2_gpu, des2_gpu = orb_gpu.detectAndCompute(gpu_initial_gray, None)

                if des1_gpu is not None and des2_gpu is not None:
                    des1 = des1_gpu.download()
                    des2 = des2_gpu.download()
                    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
                    matches = bf.match(des1, des2)
                    matches = sorted(matches, key=lambda x: x.distance)
                    good_matches = [m for m in matches[:50] if m.distance < 50]

                    if len(good_matches) > 10:
                        src_pts = np.float32([kp1_gpu[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
                        dst_pts = np.float32([kp2_gpu[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)
                        homography_matrix, mask = cv2.findHomography(dst_pts, src_pts, cv2.RANSAC, 5.0)
                        if homography_matrix is not None:
                            logger.info("Auto-alignment homography computed successfully")
                        else:
                            logger.warning("Homography computation failed")
                    else:
                        logger.warning(f"Insufficient good matches for alignment ({len(good_matches)})")
                else:
                    logger.warning("Failed to compute descriptors for alignment")
            except Exception as e:
                logger.warning(f"Auto-alignment setup failed: {e}")
            finally:
                if 'gpu_reference' in locals():
                    gpu_reference.release()
                if 'gpu_reference_gray' in locals():
                    gpu_reference_gray.release()
                if 'orb_gpu' in locals():
                    del orb_gpu

        gpu_initial.release()
        gpu_initial_gray.release()

        # Pre-load all ROI images and annotations to GPU for efficiency
        all_rois_gpu = {}  # roi_filename -> (gpu_roi_gray, folder_type, matched_annotation)
        all_annotations = []
        expected_rois = set()  # Track all expected ROIs for completion check

        # Process both GOOD and NG folders
        for i, ((roi_folder,), (json_folder,)) in enumerate(zip(roi_folders, json_folders)):
            folder_type = folder_types[i]

            logger.info(f"Accessing ROI folder ({folder_type}): {roi_folder}")
            logger.info(f"Accessing JSON folder ({folder_type}): {json_folder}")

            if not os.path.exists(roi_folder) or not os.path.exists(json_folder):
                if folder_type == "NG":
                    logger.warning(f"Folder(s) (NG) not found, skipping NG processing")
                    continue
                else:
                    logger.error(f"Folder(s) ({folder_type}) not found")
                    messagebox.showerror("Error", f"Profiling folder(s) ({folder_type}) not found")
                    raise FileNotFoundError(f"Folder(s) ({folder_type}) not found")

            json_files = [f for f in os.listdir(json_folder) if f.lower().endswith('.json')]
            if not json_files:
                if folder_type == "NG":
                    logger.warning(f"No JSON files found in {json_folder} (NG)")
                    continue
                else:
                    logger.error(f"No JSON files found in {json_folder}")
                    raise FileNotFoundError(f"No JSON files found in {json_folder}")

            for json_file in json_files:
                json_path = os.path.join(json_folder, json_file)
                try:
                    with open(json_path, 'r') as f:
                        annotations = json.load(f)
                        all_annotations.extend(annotations)
                except Exception as e:
                    logger.error(f"Failed to load JSON {json_path}: {e}")

            # Pre-load ROIs to GPU and track expected
            for roi_filename in os.listdir(roi_folder):
                if roi_filename.lower().endswith('.png'):
                    expected_rois.add(roi_filename)
                    roi_path = os.path.join(roi_folder, roi_filename)
                    roi_img = cv2.imread(roi_path, cv2.IMREAD_UNCHANGED)
                    if roi_img is None:
                        continue

                    matched_annotation = None
                    for anno in all_annotations:
                        if anno.get('ROI_image_index') == roi_filename:
                            matched_annotation = anno
                            break

                    if not matched_annotation:
                        logger.warning(f"No annotation found for ROI: {roi_filename} ({folder_type})")
                        continue

                    try:
                        roi_gray_cpu = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
                        roi_height, roi_width = roi_gray_cpu.shape
                        logger.info(f"ROI image {roi_filename} ({folder_type}) dimensions: {roi_width}x{roi_height}")

                        if roi_width > img_width or roi_height > img_height:
                            logger.warning(f"ROI {roi_filename} ({folder_type}) is larger than frame, skipping")
                            continue

                        gpu_roi_gray = cv2.cuda_GpuMat()
                        gpu_roi_gray.upload(roi_gray_cpu)

                        all_rois_gpu[roi_filename] = (gpu_roi_gray, folder_type, matched_annotation)
                    except Exception as e:
                        logger.warning(f"Failed to prepare ROI {roi_filename} ({folder_type}) for GPU: {e}")

        logger.info(f"Pre-loaded {len(all_rois_gpu)} ROI-annotation pairs. Expected ROIs: {len(expected_rois)}")

        # List to store unique successful detections (accumulate over session)
        detection_results = []
        detected_rois = set()  # Track detected ROIs to check completion

        # Live viewing loop
        frame_count = 0
        process_every_n_frames = 5  # Adjust for performance (higher = less frequent processing)
        live_window_name = "Live Annotated View"
        cv2.namedWindow(live_window_name, cv2.WINDOW_NORMAL)

        # Auto-save condition: Save when all expected ROIs detected or after max_frames
        max_frames = 300  # e.g., ~10s at 30fps; adjust as needed
        consecutive_no_new = 0
        no_new_threshold = 30  # Frames without new detections before auto-save

        logger.info("Starting live viewing. Auto-saving when all detections loaded or max frames reached. Press 'q' to quit early.")
        while frame_count < max_frames:
            ret, frame = self.cap.read()
            if not ret:
                logger.warning("Failed to grab frame")
                break

            frame_count += 1
            annotated_frame = frame.copy()
            current_gray_cpu = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2GRAY)
            gpu_current = cv2.cuda_GpuMat()
            gpu_current.upload(annotated_frame)
            gpu_current_gray = cv2.cuda.cvtColor(gpu_current, cv2.COLOR_BGR2GRAY)

            # Apply homography if available (align frame on GPU)
            if homography_matrix is not None:
                try:
                    aligned_gpu = cv2.cuda.warpPerspective(gpu_current, homography_matrix, (img_width, img_height))
                    gpu_current_gray = cv2.cuda.cvtColor(aligned_gpu, cv2.COLOR_BGR2GRAY)
                    annotated_frame = aligned_gpu.download()
                    current_gray_cpu = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2GRAY)
                except Exception as e:
                    logger.warning(f"Frame alignment failed: {e}")

            # Process matching only every N frames
            if frame_count % process_every_n_frames == 0:
                new_detections_this_frame = 0
                for roi_filename, (gpu_roi_gray, folder_type, matched_annotation) in all_rois_gpu.items():
                    try:
                        # CUDA template matching
                        gpu_result = cv2.cuda.matchTemplate(gpu_current_gray, gpu_roi_gray, cv2.TM_CCOEFF_NORMED)
                        result = gpu_result.download()

                        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
                        logger.debug(f"Frame {frame_count} - ROI {roi_filename} ({folder_type}): confidence={max_val:.2f}")

                        if max_val >= confidence_threshold:
                            x, y = max_loc
                            width = int(matched_annotation['width'])
                            height = int(matched_annotation['height'])
                            top_left = (x, y)
                            bottom_right = (x + width, y + height)

                            if 0 <= x < img_width and 0 <= y < img_height and x + width <= img_width and y + height <= img_height:
                                color = (0, 255, 0) if folder_type == "GOOD" else (0, 0, 255)
                                cv2.rectangle(annotated_frame, top_left, bottom_right, color, 2)

                                label = matched_annotation.get('serial_number', '')
                                cv2.putText(annotated_frame, f"({folder_type}) {label}", (x, y - 10),
                                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

                                # Accumulate unique detections
                                if roi_filename not in detected_rois:
                                    detected_rois.add(roi_filename)
                                    detection_results.append((label, self.tp, self.part_number, roi_filename, folder_type, MB_position))
                                    new_detections_this_frame += 1
                                    logger.info(f"New detection: {roi_filename} ({folder_type}) at frame {frame_count}")

                    except Exception as e:
                        logger.warning(f"Template matching failed for {roi_filename}: {e}")

                if new_detections_this_frame == 0:
                    consecutive_no_new += 1
                else:
                    consecutive_no_new = 0

            # Display live frame
            cv2.imshow(live_window_name, annotated_frame)

            # Check auto-save conditions
            all_detected = len(detected_rois) >= len(expected_rois)
            if all_detected or consecutive_no_new >= no_new_threshold or frame_count >= max_frames:
                logger.info(f"Auto-save triggered: All detected ({all_detected}), no new for {consecutive_no_new} frames, or max frames {frame_count}")
                break

            # Exit on 'q'
            if cv2.waitKey(1) & 0xFF == ord('q'):
                logger.info("Quit pressed, saving results.")
                break

            # Release per-frame GPU mats
            gpu_current.release()
            gpu_current_gray.release()

        cv2.destroyAllWindows()
        # Do not release self.cap here; keep it open for other uses. If needed, add self.cap.release() and self.cap = None

        # Save automatically once detections are loaded
        if detection_results:
            logger.info(f"Live session ended. Saving {len(detection_results)} detections.")

            datetime_saved = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")

            # Save to DB
            try:
                conn = pyodbc.connect(
                    f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password};TrustServerCertificate={trust_certificate};Connection timeout={connection_timeout};'
                )
                cursor = conn.cursor()
                for label, tp_val, part_number, roi_filename, detection_status, mb_pos in detection_results:
                    cursor.execute("""
                        INSERT INTO TBL_PER_LOC (Serial_number, Model, Part_number, Item_detected, Status, MB_position, Date_saved)        
                        VALUES (?, ?, ?, ?, ?, ?, ?)                
                    """, (label, tp_val, part_number, roi_filename, detection_status, mb_pos, datetime_saved))
                conn.commit()
                cursor.close()
                conn.close()
                messagebox.showinfo("Info", "Saved Data Successfully!")
                logger.info(f"Saved {len(detection_results)} detection results to DB.")
            except Exception as e:
                logger.error(f"Failed to save to DB: {e}")

            # Save last annotated frame as PNG
            output_result_path = "/home/nvidia/SHARPEYE_DATA/Result_images"
            os.makedirs(output_result_path, exist_ok=True)
            timestamp_save_result = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
            output_path = f"{output_result_path}/{self.tp}_{self.part_number}_{timestamp_save_result}.png"
            cv2.imwrite(output_path, annotated_frame)
            logger.info(f"Saved final annotated frame to {output_path}")

            # Save to Excel
            output_result_path_excel = "/home/nvidia/SHARPEYE_DATA/Excel_data"
            os.makedirs(output_result_path_excel, exist_ok=True)
            try:
                from openpyxl import Workbook
                from openpyxl.drawing.image import Image as ExcelImage
                from PIL import Image as PILImage

                wb = Workbook()
                ws = wb.active
                ws.title = "Detection Results"
                headers = ["Serial_number", "Model", "Part_number", "Item_detected", "Status", "MB_position", "Date_saved"]
                ws.append(headers)
                for col in range(1, 8):
                    ws.column_dimensions[chr(64 + col)].width = 30

                # Resize last annotated image for Excel
                temp_img_path = f"{output_result_path}/temp_annotated.png"
                pil_img = PILImage.open(output_path)
                target_size = (100, 100)
                pil_img.thumbnail(target_size, PILImage.LANCZOS)
                pil_img.save(temp_img_path, format="PNG")

                for idx, (label, tp_val, part_number, roi_filename, detection_status, mb_pos) in enumerate(detection_results, start=2):
                    ws.cell(row=idx, column=1).value = label
                    ws.cell(row=idx, column=2).value = tp_val
                    ws.cell(row=idx, column=3).value = part_number
                    ws.cell(row=idx, column=4).value = roi_filename
                    ws.cell(row=idx, column=5).value = detection_status
                    ws.cell(row=idx, column=6).value = mb_pos
                    ws.cell(row=idx, column=7).value = datetime_saved

                    img_excel = ExcelImage(temp_img_path)
                    ws.add_image(img_excel, f"H{idx}")
                    ws.row_dimensions[idx].height = 80

                timestamp_excel = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
                excel_path = f"{output_result_path_excel}/{self.tp}_{self.part_number}_Results_{timestamp_excel}.xlsx"
                wb.save(excel_path)
                wb.close()
                os.remove(temp_img_path)
                logger.info(f"Saved to Excel: {excel_path}")

                # Optional FTP upload (uncomment if needed)
                # def save_to_ftp(...): ...  # Paste the function here if using
                # save_to_ftp(excel_path, "192.168.1.152", "share", "share", "/josh/Excel_data_SE")
            except Exception as e:
                logger.error(f"Failed to save to Excel: {e}")
        else:
            logger.info("No detections in live session.")

        # Release remaining GPU resources
        for gpu_roi in all_rois_gpu.values():
            gpu_roi[0].release()

        # Optional: Display the final annotated image if needed
        # self.display_annotated_image(output_path)


    # ************************************************************************************************************************************ #

    # For Capture Profiling
    def capture_function(self):
        # get the values from serial number input and on mb position
        serial_number = self.serial_entry.get().strip()
        position_mb = self.dropdown_mb_position_profile.get()

        # initialize a list
        errors = []

        # check the serial number
        if not serial_number:
            errors.append("Serial number is required!")

        if position_mb == "---Select MB Position---":
            errors.append("Mainboard position must be selected.")

        # handle validation results
        if errors:
            if len(errors) == 2:
                messagebox.showwarning("Warning", "Both serial number and MB position is required!")
                logger.warning("Both serial number and MB position is required!")
            elif errors[0].startswith("Serial number"):
                messagebox.showinfo("Info", "Serial number is required!")
                logger.info("Serial number is missing when capturing.")
            else:
                messagebox.showinfo("Info", "MB Position is required!")
                logger.info("MB Position is missing when capturing.")
            return
        
        self.hide_video()
        time.sleep(0.5)
        self.capture.capture_photo()
        self.display_image()
        self.bind_mouse_events()

    # For NG Profiling
    def capture_function_NG(self):
        # get the values from serial number input and on mb position
        serial_number = self.serial_entry.get().strip()
        position_mb = self.dropdown_mb_position_profile_NG.get()

        # initialize a list
        errors = []

        # check the serial number
        if not serial_number:
            errors.append("Serial number is required!")

        if position_mb == "---Select MB Position---":
            errors.append("Mainboard position must be selected.")

        # handle validation results
        if errors:
            if len(errors) == 2:
                messagebox.showwarning("Warning", "Both serial number and MB position is required!")
                logger.warning("Both serial number and MB position is required!")
            elif errors[0].startswith("Serial number"):
                messagebox.showinfo("Info", "Serial number is required!")
                logger.info("Serial number is missing when capturing.")
            else:
                messagebox.showinfo("Info", "MB Position is required!")
                logger.info("MB Position is missing when capturing.")
            return

        self.hide_video()
        time.sleep(0.5)
        self.capture.capture_photo()
        self.display_image()
        self.bind_mouse_events_NG()

    # For Capture Inspection
    def capture_function_inspection(self):
        # get the values from serial number input and on mb position
        serial_number = self.serial_entry.get().strip()
        position_mb_inspect = self.dropdown_mb_position_inspect.get()

        # initialize a list
        errors_inspect = []

        # check the serial number
        if not serial_number:
            errors_inspect.append("Serial number is required!")

        if position_mb_inspect == "---Select MB Position---":
            errors_inspect.append("Mainboard position must be selected.")

            # handle validation results
        if errors_inspect:
            if len(errors_inspect) == 2:
                messagebox.showwarning("Warning", "Both serial number and MB position is required!")
                logger.warning("Both serial number and MB position is required!")
            elif errors_inspect[0].startswith("Serial number"):
                messagebox.showinfo("Info", "Serial number is required!")
                logger.info("Serial number is missing when capturing.")
            else:
                messagebox.showinfo("Info", "MB Position is required!")
                logger.info("MB Position is missing when capturing.")
            return

        self.hide_video()
        time.sleep(0.5)
        self.capture.capture_photo()
        self.display_image()
        self.unbind_mouse_events()

    def display_image(self):
        self.clear_canvas()
        self.watcher.start()

    # bind mouse events for GOOD
    def bind_mouse_events(self):
        self.canvas.bind("<Button-1>", lambda event: start_drawing(event, self))
        self.canvas.bind("<B1-Motion>", lambda event: update_rectangle(event, self))
        self.canvas.bind("<ButtonRelease-1>", lambda event: finish_drawing(event, self))

    # bind mouse events for NG
    def bind_mouse_events_NG(self):
        self.canvas.bind("<Button-1>", lambda event: start_drawing_NG(event, self))
        self.canvas.bind("<B1-Motion>", lambda event: update_rectangle_NG(event, self))
        self.canvas.bind("<ButtonRelease-1>", lambda event: finish_drawing_NG(event, self))

    # unbind the mouse events
    def unbind_mouse_events(self):
        self.canvas.unbind("<Button-1>", lambda event: start_drawing(event, self))
        self.canvas.unbind("<B1-Motion>", lambda event: update_rectangle(event, self))
        self.canvas.unbind("<ButtonRelease-1>", lambda event: finish_drawing(event, self))

        self.canvas.unbind("<Button-1>", lambda event: start_drawing_NG(event, self))
        self.canvas.unbind("<B1-Motion>", lambda event: update_rectangle_NG(event, self))
        self.canvas.unbind("<ButtonRelease-1>", lambda event: finish_drawing_NG(event, self))

    def clear_canvas(self):
        for rect in self.drawn_rectangles:
            self.canvas.delete(rect)
        self.drawn_rectangles.clear()
        logger.info("All drawn rectangles are cleared.")

    def clear_rectangles(self):
        for rect in self.drawn_rectangles:
            self.canvas.delete(rect)
        self.roi_list = []
        self.drawn_rectangles.clear()
        self.show_video()
        logger.info("All drawn rectangles are cleared.")

    def on_closing(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit the application?"):
            print("Performing clean up before closing the application.")
            self.watcher.stop()
            self.root.destroy()
        else:
            print("Exit aborted.")

    def create_folder_structure(self, base_path="/home/nvidia"):
        sharpeye_data_path = os.path.join(base_path, "SHARPEYE_DATA")

        if os.path.exists(sharpeye_data_path):
            logger.info(f"SHARPEYE_DATA exist at {sharpeye_data_path}. Skipping creation.")
        else:
            os.makedirs(sharpeye_data_path, exist_ok=True)
            logger.info(f"Created SHARPEYE_DATA at {sharpeye_data_path}")

        subfolders = [
            "Excel_data",
            "Profiling_data",
            "Profiling_data_NG",
            "Resize_images",
            "Result_images",
            "Captured_images",
            "Assets"
        ]

        for subfolder in subfolders:
            subfolder_path = os.path.join(sharpeye_data_path, subfolder)
            if os.path.exists(subfolder_path):
                logger.info(f"Subfolders already exist: {subfolder_path}")
            else:
                os.makedirs(subfolder_path, exist_ok=True)
                logger.info(f"Created subfolder: {subfolder_path}")

    def run(self):
        self.create_folder_structure()
        self.root.mainloop()

if __name__ == "__main__":
    root = ctk.CTk()
    app = MainApplication(root)
    global_function_tp_pn(app)
    app.run()